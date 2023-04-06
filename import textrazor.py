import textrazor
from textrazor import TextRazor
import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import config
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
from openpyxl import Workbook
from datetime import datetime


TEXTRAZOR_API_KEY = "37dXXXXXXXXXXXXXXXXYYYYYYYYYYYHHHHHHHHHHHGGGGGGGGGG"
textrazor.api_key = TEXTRAZOR_API_KEY #my real API masked
client = textrazor.TextRazor()
# textrazor.api_key = ''

# Specify the keyword
keyword = "hardwood floor refinishing los angeles"
urls = [
    '',
    '',
    '',
    '',
    ''
]

#def sanitize_filename(filename):
    # Remove characters that are not allowed in filenames on Windows
   # return re.sub(r'[<>:"/\\|?*.]', '_', filename)

# overcoming the str problem
def custom_code(obj):
    if isinstance(obj, float) or isinstance(obj, int):
        return "{:.2f}".format(obj)
    else:
        return "{}".format(obj)

wb = Workbook()
ws = wb.active

def get_row_data_for_data_type(item, data_type):
    if data_type == 'entities':
        row_data = [
            item.get('id', ''),
            item.get('wikipedia_link', ''),
            item.get('matched_text', ''),
            item.get('starting_position', ''),
            ", ".join(item.get('freebase_types', [])),
            item.get('dbpedia_types', ''),
            item.get('relevance_score', ''),
            item.get('confidence_score', ''),
            ", ".join([word.token for word in item.get('matched_words', [])]),
        ]
    if data_type == 'entailments':
        row_data = [
            item.id,
            ", ".join(str(pos) for pos in item.matched_positions),
            item.prior_score,
            item.context_score,
            item.score,
            ", ".join(str(word) for word in item.get('matched_words', [])),
            item.entailed_word,
        ]
    elif data_type == 'topics':
        row_data = [
             item.id,
            item.score,
            item.wikipedia_link,
            item.wikidata_id,
            item.label,
        ]
    elif data_type == 'categories':
         row_data = [
            item.query,
            item.label,
            item.classifier_id,
        ]
    elif data_type == 'property_properties':
        row_data = [
            item.id,
            ", ".join(item.get('property_words', [])),
            ", ".join(item.get('predicate_words', [])),
        ]
    elif data_type == 'noun_phrases':
        row_data = [
             item.id,
             item.word_position,
             item.words,
        ]
    elif data_type == 'relations':
        row_data = [
            item.id,
            item.params,
            item.predicate_words,
        ]
    elif data_type =='relation_params':
        row_data = [
            item.entities,
             item.param_words,
             item.relation,
              item.relation_parent,
        ]
    elif data_type == 'sentences':
        row_data = [
            item.root_word,
            item.words,
        ]
    elif isinstance(item, TextRazor) and data_type == 'words':
        row_data = []
        if item.parent is not None:
            row_data = [
                item.parent.id,
                item.relation_to_parent,
            ]
        row_data += [
            item.children,
            item.entailments,
            item.entities,
            item.part_of_speech,
            item.lemma,
            item.noun_phrases,
            item.relations,
            item.prelation_params,
            item.senses,
            item.spelling_suggestions,
            item.root_word,
            item.words,
        ]
        
    else:
        row_data = []
    return row_data

def get_headers_for_data_type(data_type):
    if data_type == 'entities':
        return ['ID', 'Wikipedia Link', 'Matched Text', 'Starting Position', 'Freebase Types', 'DBPedia Types', 'Relevance Score', 'Confidence Score', 'Matched Words']
    if data_type == 'entailments':
        return ['ID', 'Matched Positions', 'Prior Score', 'Context Score', 'Score', 'Matched Words', 'Entailed Word']
    if data_type == 'topics':
        return ['ID', 'Score', 'Wikipedia Link', 'Wikidata ID', 'Label']
    if data_type == 'categories':
        return ['Query', 'Label', 'Classifier ID']
    if data_type == 'property_properties':
        return ['ID', 'Property Words', 'Predicate Words']
    if data_type == 'noun_phrases':
        return ['ID', 'Word Position', 'Words']
    if data_type == 'relations':
        return ['ID', 'Params', 'Predicate Words']
    if data_type == 'relation_params':
        return ['Entities', 'Param Words', 'Relation', 'Relation Parent']
    if data_type == 'sentences':
        return ['Root Word', 'Words']
    if data_type == 'words':
        return ['Parent', 'Relation to Parent', 'Children', 'Entailments', 'Entities', 'Part of Speech', 'Lemma', 'Noun Phrases', 'Relations', 'Relation Params', 'Senses', 'Spelling Suggestions']
    else:
        return []
    

# Scrape sitemap and retrieve new pages/posts
def crawl_page(url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'}
    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        print(f"Failed to crawl URL '{url}' with status code {response.status_code}")
        return None

    soup = BeautifulSoup(response.content, "html.parser")

    # Extract the text wrapped in the specified tags
    relevant_tags = ['h1', 'h2', 'h3', 'h4', 'p', 'b', 'strong', 'i', 'ul', 'li']
    text_parts = []
    for tag in soup.find_all(relevant_tags):
        text_parts.append(tag.get_text())

    # Extract the anchor text without the hyperlink URL
    for a in soup.find_all("a"):
        anchor_text = a.get_text(strip=True)
        if anchor_text:
            text_parts.append(anchor_text)

    # Combine the extracted text parts
    page_text = ' '.join(text_parts)
    return page_text


def extract_all_data(text, client):
    client.set_extractors(["entities", " dependency-trees", "phrases" "topics", "relations", "entailments", "words"])
    client.set_language_override("eng")
    client.set_classifiers(["textrazor_newscodes"])
    chunk_size = 10000
    text_chunks = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]

    extracted_data = {
        "entities": [],
        "entailments": [],
        "topics": [],
        "categories": [],
        "property_properties": [],
        "noun_phrases": [],
        "relations": [],
        "relation_params": [],
        "sentences": [],
        "words": [],
    }

    for chunk in text_chunks:
        if chunk:
            print(f"Analyzing chunk: {chunk[:100]}...")
            response = client.analyze(chunk)
        for entity in response.entities():
            entity_data = {
                "id": entity.id,
                "wikipedia_link": entity.wikipedia_link,
                "matched_text": entity.matched_text,
                "starting_position": entity.starting_position,
                "freebase_types": entity.freebase_types,
                "dbpedia_types": entity.dbpedia_types,  # added this line
                "relevance_score": entity.relevance_score,
                "confidence_score": entity.confidence_score,
                "matched_words": entity.matched_words,  # added this line
            }
            extracted_data["entities"].append(entity_data)

        for entailment in response.entailments():
            entailment_data = {
                "id": entailment.id,  # added this line
                "context_score": entailment.context_score,
                "entailed_word": entailment.entailed_word,
                "matched_positions": entailment.matched_positions,
                "matched_words": entailment.matched_words,
                "prior_score": entailment.prior_score,
                "score": entailment.score,
            }
            extracted_data["entailments"].append(entailment_data)

        for topic in response.topics():
            topic_data = {
                "id": topic.id,
                "score": topic.score,
                "wikipedia_link": topic.wikipedia_link,
                "wikidata_id": topic.wikidata_id,
                "label": topic.label,
            }
            extracted_data["topics"].append(topic_data)

        for category in response.categories():
            category_data = {
                "score": category.score,  # changed "label" to "query" to score
                "label": category.label,
                "classifier_id": category.classifier_id,
            }
            extracted_data["categories"].append(category_data)


        for property in response.properties():
            property_data = {
                "id": property.id,  # added this line
                "property_words": property.property_words,
                "predicate_words": property.predicate_words,
            }
            extracted_data["property_properties"].append(property_data)  # changed "properties" to "property_properties"
       
        for phrase in response.noun_phrases():
            phrase_data = {
                "id": phrase.id,  # added this line
                "word_positions": phrase.word_positions,
                "words": phrase.words
            }
            extracted_data["noun_phrases"].append(phrase_data)
       
        for relation in response.relations():
            relation_data = {
                "id": relation.id,  # added this line
                "params": relation.params,
                "predicate_words": relation.predicate_words,
            }
            extracted_data["relations"].append(relation_data)
                
        for relation in response.relations():
            for param in relation.params:
                relation_param_data = {
                    "entities": param.entities,
                    "param_words": param.param_words,
                    "relation": param.relation,
                    "relation_parent": param.relation_parent
            }
            extracted_data["relation_params"].append(relation_param_data)

        for sentence in response.sentences():
            sentence_data = {
                "root_word": sentence.root_word,
                "words": sentence.words
            }
            extracted_data["sentences"].append(sentence_data)

        for word in response.words():
            word_data = {
                "children": word.children,
                "entailments": word.entailments,
                "entities": word.entities,
                "part_of_speech": word.part_of_speech,
                "lemma": word.lemma,
                "noun_phrases": word.noun_phrases,
                "parent": word.parent,
                "relation_to_parent": word.relation_to_parent,  # added this line
                "relations": word.relations,
                "relation_params": word.relation_params,  # added this line
                "senses": word.senses,
                "spelling_suggestions": word.spelling_suggestions
            }
            extracted_data["words"].append(word_data)


        return extracted_data

data_types = ['entities', 'entailments', 'topics', 'categories', 'property_properties', 'noun_phrases', 'relations', 'relation_params', 'sentences', 'words']

for url in urls:
    text = crawl_page(url)
    if text is None:
        continue

    extracted_data = extract_all_data(text, client)

    # Create a new sheet for each URL
    url_ws = wb.create_sheet(title=urlparse(url).netloc)

    # Iterate over data types
    for data_type in data_types:
        # Create a new sheet for each data type for each URL
        ws = wb.create_sheet(title=f"{urlparse(url).netloc}_{data_type}")

        # Add headers to the sheet
        headers = get_headers_for_data_type(data_type)
        ws.append(headers)

        # Add data to the sheet
        for item in extracted_data[data_type]:
            row_data = get_row_data_for_data_type(item, data_type)
            ws.append(row_data)

        for data_type in extracted_data.keys():
            headers = get_headers_for_data_type(data_type)
        for item in extracted_data[data_type]:
            row_data = get_row_data_for_data_type(item, data_type)
            ws.append(row_data)
            
# Create a folder named 'results' in the same directory as the script
output_folder = 'results'
os.makedirs(output_folder, exist_ok=True)

# Remove the default sheet that is created when initializing the workbook
wb.remove(wb.active)

def analyze_entities(text):
    extracted_data = extract_all_data(text)
    return extracted_data


            
# Generate a timestamp and add it to the filename
timestamp = datetime.now().strftime("%m-%d-%Y_%I-%M-%p")
filename = f'text_razor_{keyword}_{timestamp}.xlsx'

output_path = os.path.join(output_folder, filename)
# Save the modified workbook
wb.save(output_path)
